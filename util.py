from os import path
import numpy as np
from numpy import core
import pandas as pd
import csv
import os
import string
from PySide2.QtCore import *
from PySide2.QtGui import *
from PySide2.QtWidgets import *
import pathlib
import xlsxwriter
import xlwt
import openpyxl

def pop_window_text(title, text):
    msg = QMessageBox()
    msg.setWindowTitle(title)
    msg.setText(text)
    msg.setIcon(QMessageBox.Critical)
    msg.setStandardButtons(QMessageBox.Cancel|QMessageBox.Retry|QMessageBox.Ignore)
    msg.setDefaultButton(QMessageBox.Cancel)
    x = msg.exec_()

def pop_window_input(title, text):
    #TBD
    return 0

def read_from_window(window, single):
    data = []
    if single:
        for row in range(window.rowCount()):
            item = window.item(row, 0)
            if item is not None:
                data.append(item.text())
            else:
                data.append('')
        return data
    
    else:
        for row in range(window.rowCount()):
            rowdata = []
            for column in range(window.columnCount()):
                item = window.item(row, column)
                if item is not None:
                    rowdata.append(item.text())
                else:
                    rowdata.append('')
            data.append(rowdata)
        return data

def write_to_csv(window, path):
    with open(path, 'w') as stream:
        writer = csv.writer(stream)
        for row in range(window.rowCount()):
            rowdata = []
            for column in range(window.columnCount()):
                item = window.item(row, column)
                if item is not None:
                    rowdata.append(item.text())
                else:
                    rowdata.append('')
            writer.writerow(rowdata)

def write_to_excel(window, path, sheetname = 'theory'):

    if not os.path.isfile(path):
        workbook= openpyxl.Workbook()
    else:
        workbook = openpyxl.load_workbook(path)
    
    worksheet = workbook.create_sheet()
    worksheet.title = sheetname

    # summary workbook
    sum_path = './sum.xlsx'
    if not os.path.isfile(sum_path):
        workbook_sum= openpyxl.Workbook()
    else:
        workbook_sum = openpyxl.load_workbook(path)

    worksheet_sum = workbook_sum.create_sheet()
    worksheet_sum.title = sheetname

    for row in range(1, 1+window.rowCount()):
        for column in range(1, 1+window.columnCount()):
            item = window.item(row-1, column-1)
            worksheet.cell(row, column, item.text())
            worksheet_sum.cell(row, column, item.text())
    
    workbook.save(path)
    workbook_sum.save(sum_path)

def clear_window(window):
    window.setRowCount(0)
    window.setColumnCount(0)
    window.clearContents()

def display_in_window(input, window, coordinate_names):
    clear_window(window)
    rows, cols = input.shape[0], input.shape[1]
    print(rows, cols)
    header = ['coordinate'] + input.columns.values.tolist()
    print(header)
    if(rows != len(coordinate_names)):
        pop_window_text('Warning', 'Number of coordinates cannot match')
        return
    
    window.setColumnCount(cols+1)
    window.setRowCount(rows+1)
    # window.setHorizontalHeaderLabels(header)

    for i in range(rows):
        newItem = QTableWidgetItem(str(coordinate_names[i]))
        newItem.setTextAlignment(Qt.AlignHCenter|Qt.AlignVCenter)
        window.setItem(i+1, 0, newItem)

    for j in range(cols+1):
        newItem = QTableWidgetItem(str(header[j]))
        newItem.setTextAlignment(Qt.AlignHCenter|Qt.AlignVCenter)
        window.setItem(0, j, newItem)

    for i in range(rows):
        for j in range(cols):
            newItem = QTableWidgetItem(str(input.iloc[i][j]))
            newItem.setTextAlignment(Qt.AlignHCenter|Qt.AlignVCenter)
            window.setItem(i+1, j+1, newItem)
    
    window.resizeColumnsToContents()

def read_from_csv(window, path, coordinate_names):
    input_table = pd.read_csv(path)
    display_in_window(input_table, window, coordinate_names)

def read_from_excel(window, path, sheet, coordinate_names):
    input_table = pd.read_excel(path, sheet)
    display_in_window(input_table, window, coordinate_names)