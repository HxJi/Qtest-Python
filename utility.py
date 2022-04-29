import collections
from os import path
import numpy as np
from numpy import core, double
import pandas as pd
import csv
import os
import string
from PySide2.QtCore import *
from PySide2.QtGui import *
from PySide2.QtWidgets import *
import pathlib
import xlsxwriter
import xlwt
import openpyxl
import string
from itertools import combinations

def generate_coordinate(num):
    elements = list(string.ascii_uppercase)[:num]
    return combinations(elements, 2)    

def display_coordinate(window, coo, num_coo):
    for i in range(num_coo):
        newitem = QTableWidgetItem(str(coo[i]))
        newitem.setTextAlignment(Qt.AlignHCenter|Qt.AlignVCenter)
        window.setItem(0,i,newitem)
    window.setHorizontalHeaderLabels(['Coordinate'])
    window.resizeColumnsToContents()

def add_row_in_window(window, count):
    if(window.rowCount() == 0):
        window.setColumnCount(1)
    window.setRowCount(window.rowCount()+count)

def remove_row_in_window(window, rows):
    for row in rows:
        window.removeRow(row)
        print(row)
    window.resizeColumnsToContents()

def clear_window(window):
    window.setRowCount(0)
    window.setColumnCount(0)
    window.clearContents()

def write_to_csv(window, path, header):
    with open(path, 'w') as stream:
        writer = csv.writer(stream)
        if header != None:
            writer.writerow(header)
        for row in range(window.rowCount()):
            rowdata = []
            for column in range(window.columnCount()):
                item = window.item(row, column)
                if item is not None:
                    rowdata.append(item.text())
                else:
                    rowdata.append('')
            writer.writerow(rowdata)

def write_to_excel(window, path, sheetname = 'theory'):

    if not os.path.isfile(path):
        workbook= openpyxl.Workbook()
    else:
        workbook = openpyxl.load_workbook(path)
    
    worksheet = workbook.create_sheet()
    worksheet.title = sheetname

    # summary workbook
    sum_path = './summary.xlsx'
    if not os.path.isfile(sum_path):
        workbook_sum= openpyxl.Workbook()
    else:
        workbook_sum = openpyxl.load_workbook(path)

    worksheet_sum = workbook_sum.create_sheet()
    worksheet_sum.title = sheetname

    for row in range(1, 1+window.rowCount()):
        for column in range(1, 1+window.columnCount()):
            item = window.item(row-1, column-1)
            worksheet.cell(row, column, item.text())
            worksheet_sum.cell(row, column, item.text())
    
    workbook.save(path)
    workbook_sum.save(sum_path)

def pop_window_text(title, text):
    msg = QMessageBox()
    msg.setWindowTitle(title)
    msg.setText(text)
    msg.setIcon(QMessageBox.Critical)
    msg.setStandardButtons(QMessageBox.Cancel|QMessageBox.Retry|QMessageBox.Ignore)
    msg.setDefaultButton(QMessageBox.Cancel)
    x = msg.exec_()

def display_theory(window, input, coo_num):
    rows, cols = input.shape[0], input.shape[1]
    header = input.columns.values.tolist()
    print(header)
    if(rows != coo_num):
        pop_window_text('Warning', 'Number of coordinates cannot match')
        print(rows, coo_num)
        return
    
    window.setColumnCount(cols)

    for j in range(cols):
        newItem = QTableWidgetItem(str(header[j]))
        newItem.setTextAlignment(Qt.AlignHCenter|Qt.AlignVCenter)
        window.setItem(0, j, newItem)

    for i in range(rows):
        for j in range(cols):
            newItem = QTableWidgetItem(str(input.iloc[i][j]))
            newItem.setTextAlignment(Qt.AlignHCenter|Qt.AlignVCenter)
            window.setItem(i+1, j, newItem)
    
    window.resizeColumnsToContents()


def check_col_values(window, lock):
    col_count = window.columnCount()
    if lock:
        row_num = window.rowCount()
        col_num = window.columnCount()
        start = 0
    else:
        row_num = window.rowCount()
        col_num = window.columnCount()
        start = 1
    
    for column in range(col_num):
        for row in range(start, row_num):
            item = window.item(row, column)
            value = float(item.text())
            if value > 1 or value < 0.5:
                pop_window_text('Error', 'Vertex Value wrong col:%d, row%d'%(column, row))
                return
    return 0

def order_intensity_cal(window, lock):
    col_count = window.columnCount()
    row_num = window.rowCount()
    col_num = window.columnCount()
    if lock:
        start = 0
    else:
        start = 1

    index_order = []
    for column in range(col_num):
        intensity = []
        for row in range(start, row_num):
            item = window.item(row, column)
            value = float(item.text())
            intensity.append(value)
        
        s = np.array(intensity)
        index_order.append(np.argsort(s))

    return index_order

def generate_p_string(list, name_list):
    p_order_string = '0<='
    for i in list:
        p_order_string = p_order_string + '%s<='%(name_list[i])
    p_order_string += '1'
    return p_order_string


def getInt(window, title, label):
    int, okPressed = QInputDialog.getInt(window,title, label)
    return int

def getDouble(window, title, label):
    double, okPressed = QInputDialog.getDouble(window,title, label)
    return double

def getText(window, title, label):
    text, okPressed = QInputDialog.getText(window,title, label)
    return text